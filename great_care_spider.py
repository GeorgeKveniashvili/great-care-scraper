import openpyxl
from scrapy.spiders import CrawlSpider, Rule
from scrapy.linkextractors import LinkExtractor

GREAT_CARE_URL = 'https://www.iwantgreatcare.org/search'
ALLOWED_EXTRACTOR = ['/doctors/', '/optometrists/', '/nurses/', '/dentists/', '/physiotherapists/', '/dietitians/', '/occupationaltherapists/']
TITLE_XPATH = '//*[@id="entity-name-score-container"]/h1'
SPECIALISES_XPATHES = ['//*[@id="specialies-container"]/div/ul', '//*[@id="information"]/div[2]/div[2]/div/ul', 
                       '//*[@id="information"]/div[1]/div[2]/div/ul']
WORKS_XPATHES = ['//*[@id="works-at-container"]/div/ul', '//*[@id="information"]/div[2]/div[1]/div/ul', 
                 '//*[@id="information"]/div[1]/div[1]/div/ul', '/html/body/div[1]/main/div/span/div[2]/div/div[2]/div[1]/ul',
                 '/html/body/div[1]/main/div/span/div[2]/div/div[2]/div[2]/div[1]/div/ul']
PROFILE_XPATHES = ['/html/body/div[1]/main/div/span/div[2]/div/div[2]/div[4]/div/div/p', 
                   '/html/body/div[1]/main/div/span/div[2]/div/div[2]/div[2]/div/div/p']
FILE_PATH = "DoctorInfo.xlsx"
FILE_INDEX = 2

class GreatCareSpider(CrawlSpider):
    name = 'greatcarespider'
    allowed_domains = ['iwantgreatcare.org']
    start_urls = [GREAT_CARE_URL]

    rules = (
        # Extract and follow all links!
        Rule(LinkExtractor(allow=ALLOWED_EXTRACTOR), callback='parse_item', follow=True),
    )

    def parse_item(self, response):
        global FILE_INDEX
        
        for title in response.xpath(TITLE_XPATH):
            full_name_formated = split_name(title.css('::text').get())
            name = full_name_formated[0]
            surname = full_name_formated[1]

            specialises = extract_multiple_selectors(response, SPECIALISES_XPATHES, 'li::text', surname)
            profile = extract_multiple_selectors(response, PROFILE_XPATHES, '::text', surname)
            works = extract_multiple_selectors(response, WORKS_XPATHES, 'a::text', surname)

            write_file(FILE_INDEX, name, surname, specialises, profile, works, response.url)
            FILE_INDEX += 1
        self.log('crawling'.format(response.url))

def extract_multiple_selectors(response, xpathes, text_selector, name):
    for xpath in xpathes:
        selector = response.xpath(xpath)
        text_list = selector.css(text_selector).getall()
        if len(text_list) > 0:
            print("The selector is good: "+ str(selector)+"name: "+name)
            break
        print("The selector is None: "+ str(selector)+"name: "+name)
    text = ""
    for item in text_list:
        text += str(item)
    return text


def write_file(index, name, surname, specialises, profile, works, domain):
    wb = openpyxl.load_workbook(FILE_PATH)
    ws = wb.get_sheet_by_name('Sheet')

    ws.cell(row=index, column=1).value = name
    ws.cell(row=index, column=2).value = surname
    ws.cell(row=index, column=3).value = specialises
    ws.cell(row=index, column=4).value = profile
    ws.cell(row=index, column=5).value = works
    ws.cell(row=index, column=6).value = domain

    wb.save(FILE_PATH)
    wb.close()

def split_name(full_name):

    if(len(full_name.split(' ')) > 2):
        name = ' '.join(full_name.split(' ', 2)[:2])
        surname = ' '.join(full_name.split(' ', 2)[2:])
        return [name, surname]
    name = full_name.split(' ', 1)[0]
    surname = full_name.split(' ', 1)[1]
    return [name, surname]